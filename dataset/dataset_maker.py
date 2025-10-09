import os
import json
import random

from openpyxl import load_workbook

#ORGデータセットを参照し、Train/valid用のデータセットに再分配
class DatasetMaker:
    def __init__(self, makedate, header_row=1, train_ratio=0.8, seed=42):
        self.header_row = header_row
        self.train_ratio = train_ratio
        self.seed = seed
        self.Reffolder_dir = os.path.join(os.path.dirname(__file__), "ORG")
        self.jsonl_dir = os.path.join(os.path.dirname(__file__), makedate)
        self.dataset_dir = os.path.join(self.jsonl_dir, f"seed{seed}_train{int(train_ratio*100)}_valid{int((1-train_ratio)*100)}")
        # 既存のデータセットがあるか確認
        self.Dir_exists = False
        if os.path.exists(self.dataset_dir):
            self.Dir_exists = True
        # 出力ディレクトリ作成(なければ)
        os.makedirs(self.jsonl_dir, exist_ok=True)
        os.makedirs(self.dataset_dir, exist_ok=True)

    def main(self):
        # 既存のデータセットがあれば使用
        if self.Dir_exists:
            print(f"既存のファイルを使用します: {self.dataset_dir}")
            return

        # 既存のデータセットが無ければ新規にデータセットを作成
        print(f"新規にデータセットを作成します: {self.dataset_dir}")
        #QA形式のデータセットを作成
        QAjsonl_path = self.ORG_to_dataset(ReffileType="QA")
        #Text形式のデータセットを作成
        Textjsonl_path = self.ORG_to_dataset(ReffileType="Text")
        
        #作成したデータセット一覧
        jsonl_paths = [QAjsonl_path, Textjsonl_path]

        #QA形式とText形式のデータセットを統合(参考用)
        self.merge_selected_jsonl_files(jsonl_paths)
        
        #QA形式とText形式から任意の割合で分配したtrain/valid用データセットを作成
        self.split_jsonl_files(jsonl_paths)

    def ORG_to_dataset(self, ReffileType="QA"):
        #参照先確認
        Reffolder_dir = os.path.join(self.Reffolder_dir, ReffileType)
        #出力ファイルパス
        output_path = os.path.join(self.jsonl_dir, f"{ReffileType}_dataset.jsonl")
        # JSONLファイル作成
        with open(output_path, 'w', encoding='utf-8') as jsonl_file:
            # フォルダ内のすべてのExcelファイルを処理
            for filename in os.listdir(Reffolder_dir):
                if filename.endswith((".xlsx",".xls",".xlsm")):
                    file_path = os.path.join(Reffolder_dir, filename)
                    wb = load_workbook(file_path, data_only=True)  # マクロは無視、値のみ取得
                    # すべてのシートを処理
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]

                        # ---- 見出し行から列を特定 ----
                        header_cells = ws[self.header_row]
                        if ReffileType == "QA":
                            input_col = None
                            output_col = None
                        elif ReffileType == "Text":
                            text_col = None
                        
                        for cell in header_cells:
                            if cell.value is None:
                                continue
                            header_text = str(cell.value).lower()

                            if ReffileType == "QA":
                                if any(keyword in header_text for keyword in ["質問", "question", "prompt", "input"]):
                                    input_col = cell.column_letter
                                elif any(keyword in header_text for keyword in ["回答", "answer", "response", "output"]):
                                    output_col = cell.column_letter
                            elif ReffileType == "Text":
                                if any(keyword in header_text for keyword in ["本文", "原文", "text"]):
                                    text_col = cell.column_letter
                        
                        #---- 必要な列が見つからない場合はスキップ ----
                        if ReffileType == "QA":
                            if not input_col or not output_col:
                                print(f"[警告] {filename} の {sheet_name} で適切な列が見つかりませんでした。")
                                continue
                        elif ReffileType == "Text":
                            if not text_col:
                                print(f"[警告] {filename} の {sheet_name} で適切な列が見つかりませんでした。")
                                continue

                        # ---- データ行を読み込み ----
                        row = self.header_row + 1
                        while True:
                            if ReffileType == "QA":
                                cell_input = ws[f'{input_col}{row}'].value
                                cell_output = ws[f'{output_col}{row}'].value
                                # 終了条件（両方空なら終了）
                                if cell_input is None and cell_output is None:
                                    break
                                record = {
                                    "input": cell_input if cell_input is not None else "",
                                    "output": cell_output if cell_output is not None else "",
                                }
                            elif ReffileType == "Text":
                                cell_text = ws[f'{text_col}{row}'].value
                                # 終了条件（空なら終了）
                                if cell_text is None:
                                    break
                                record = {
                                    "text": cell_text if cell_text is not None else "",
                                }

                            # JSONLに書き込み
                            jsonl_file.write(json.dumps(record, ensure_ascii=False) + '\n')
                            row += 1

        print(f"完了：{output_path} に保存されました。")
        return output_path

    def merge_selected_jsonl_files(self, jsonl_paths):
        # 出力ファイルパス
        output_path = os.path.join(self.jsonl_dir, "All_dataset.jsonl")
        # JSONLファイルを結合
        with open(output_path, "w", encoding="utf-8") as outfile:
            for file_path in jsonl_paths:
                if not os.path.exists(file_path):
                    print(f"⚠️ ファイルが見つかりません: {file_path}")
                    continue

                with open(file_path, "r", encoding="utf-8") as infile:
                    for line_num, line in enumerate(infile, start=1):
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            json.loads(line)  # JSONとして有効かチェック
                            outfile.write(line + "\n")
                        except json.JSONDecodeError:
                            print(f"⚠️ 無効なJSON行（{file_path}:{line_num}）をスキップ")
                            continue

        print(f"✅ {len(jsonl_paths)} 件のファイルを結合しました → {output_path}")

    def split_jsonl_files(self, jsonl_paths):
        # 乱数シード設定
        random.seed(self.seed)
        # 出力ディレクトリ作成
        train_output_path = os.path.join(self.dataset_dir, "train.jsonl")
        valid_output_path = os.path.join(self.dataset_dir, "valid.jsonl")

        # 分割して書き込み
        total_train, total_valid = 0, 0
        with open(train_output_path, "w", encoding="utf-8") as train_file, \
            open(valid_output_path, "w", encoding="utf-8") as valid_file:
            # 各ファイルを処理
            for file_path in jsonl_paths:
                if not os.path.exists(file_path):
                    print(f"⚠️ ファイルが見つかりません: {file_path}")
                    continue

                # 各ファイルのデータを読み込み
                with open(file_path, "r", encoding="utf-8") as f:
                    lines = [line.strip() for line in f if line.strip()]

                # JSONとして有効なものだけ抽出
                valid_lines = []
                for i, line in enumerate(lines, start=1):
                    try:
                        json.loads(line)
                        valid_lines.append(line)
                    except json.JSONDecodeError:
                        print(f"⚠️ 無効なJSON行をスキップ: {file_path}:{i}")

                # データをシャッフルして分割
                random.shuffle(valid_lines)
                split_index = int(len(valid_lines) * self.train_ratio)
                train_lines = valid_lines[:split_index]
                valid_lines = valid_lines[split_index:]

                # ファイルに書き込み
                for line in train_lines:
                    train_file.write(line + "\n")
                for line in valid_lines:
                    valid_file.write(line + "\n")

                total_train += len(train_lines)
                total_valid += len(valid_lines)

                print(f"✅ {os.path.basename(file_path)} を分割しました "
                    f"(train: {len(train_lines)}, valid: {len(valid_lines)})")

        print(f"\n📊 すべてのファイルを分割完了")
        print(f"Train 合計: {total_train}")
        print(f"Valid 合計: {total_valid}")
        print(f"出力先: {self.dataset_dir}")

if __name__ == "__main__":
    import datetime
    makedate = datetime.datetime.now().strftime("%Y%m%d") 

    DatasetMaker(makedate=makedate).main()